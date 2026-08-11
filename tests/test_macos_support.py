import json
import tempfile
import unittest
from pathlib import Path
from types import SimpleNamespace
from unittest import mock

from openpyxl import Workbook, load_workbook

import SMWStreamTracker_MARIO_UI_STATS_CHARTS_MARIO_TIGHTER as tracker


class MacPlatformSupportTests(unittest.TestCase):
    def test_native_application_data_directories(self):
        home = Path("/Users/mario")
        self.assertEqual(
            tracker.platform_application_data_directory(
                "darwin",
                home,
                {},
            ),
            home
            / "Library"
            / "Application Support"
            / "SMWStreamTracker",
        )
        self.assertEqual(
            tracker.platform_application_data_directory(
                "win32",
                Path("C:/Users/Mario"),
                {"LOCALAPPDATA": "C:/Users/Mario/AppData/Local"},
            ),
            Path("C:/Users/Mario/AppData/Local/SMWStreamTracker"),
        )

    def test_connection_services_accept_windows_and_mac_names(self):
        for name in ("sni", "sni.exe"):
            self.assertEqual(tracker.connection_service_kind(name), "SNI")
        for name in ("QUsb2Snes", "QUsb2Snes.exe"):
            self.assertEqual(
                tracker.connection_service_kind(name),
                "QUsb2Snes",
            )

    def test_retroarch_core_download_matches_mac_architecture(self):
        self.assertIn(
            "/arm64/",
            tracker.retroarch_core_download_url("arm64"),
        )
        self.assertIn(
            "/x86_64/",
            tracker.retroarch_core_download_url("x86_64"),
        )
        self.assertTrue(
            tracker.retroarch_core_filename("darwin").endswith(".dylib")
        )

    def test_mac_retroarch_core_uses_the_explicit_config_tree(self):
        executable = Path(
            "/Users/mario/Applications/RetroArch.app/Contents/MacOS/RetroArch"
        )
        with (
            mock.patch.object(tracker, "IS_MACOS", True),
            mock.patch.object(
                tracker.Path,
                "home",
                return_value=Path("/Users/mario"),
            ),
        ):
            self.assertEqual(
                tracker.retroarch_core_directory(executable),
                Path("/Users/mario/.config/retroarch/cores"),
            )

    def test_retroarch_tracker_config_exposes_core_and_network_settings(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            config_path = root / "retroarch.cfg"
            core_directory = root / "cores"
            tracker.write_retroarch_tracker_settings(
                config_path,
                core_directory=core_directory,
            )
            text = config_path.read_text(encoding="utf-8")

        self.assertIn('network_cmd_enable = "true"', text)
        self.assertIn('network_cmd_port = "55355"', text)
        self.assertIn(
            "libretro_directory = " + json.dumps(str(core_directory.resolve())),
            text,
        )
        self.assertIn(
            "libretro_info_path = " + json.dumps(str(core_directory.resolve())),
            text,
        )
        self.assertIn('menu_show_load_core = "true"', text)
        self.assertIn('menu_show_advanced_settings = "true"', text)

    def test_every_optional_companion_has_an_official_mac_package(self):
        self.assertTrue(
            tracker.MACOS_SNI_DOWNLOAD_URL.endswith(
                "darwin-universal.tar.gz"
            )
        )
        self.assertTrue(
            tracker.MACOS_QUSB2SNES_DOWNLOAD_URL.endswith(".dmg")
        )
        self.assertTrue(
            tracker.MACOS_RETROARCH_DOWNLOAD_URL.endswith(".dmg")
        )
        for checksum in (
            tracker.MACOS_SNI_DOWNLOAD_SHA256,
            tracker.MACOS_QUSB2SNES_DOWNLOAD_SHA256,
        ):
            self.assertRegex(checksum, r"^[0-9a-f]{64}$")

    def test_mac_build_pipeline_targets_apple_silicon_and_intel(self):
        project_root = Path(__file__).resolve().parents[1]
        workflow = (
            project_root / ".github" / "workflows" / "build-macos.yml"
        ).read_text(encoding="utf-8")
        build_script = (
            project_root / "release" / "build_macos_release.sh"
        ).read_text(encoding="utf-8")
        requirements = (
            project_root / "release" / "requirements-macos.txt"
        ).read_text(encoding="utf-8")
        launcher = (
            project_root / "SMWStreamTrackerLauncher.py"
        ).read_text(encoding="utf-8")
        spec = (
            project_root / "SMWStreamTracker-macOS.spec"
        ).read_text(encoding="utf-8")

        self.assertIn("macos-15", workflow)
        self.assertIn("macos-15-intel", workflow)
        self.assertIn("arm64", workflow)
        self.assertIn("x86_64", workflow)
        self.assertIn('tags:\n      - "v*"', workflow)
        self.assertIn("gh release upload", workflow)
        self.assertIn("hdiutil create", build_script)
        self.assertIn("codesign --verify", build_script)
        self.assertIn("notarytool submit", build_script)
        self.assertIn("pyobjc", requirements)
        self.assertIn("certifi==2026.7.22", requirements)
        self.assertNotIn("pystray", requirements)
        self.assertIn("configure_platform_runtime(tracker)", launcher)
        self.assertIn("configure_secure_networking()", launcher)
        self.assertIn("run_macos_network_check()", launcher)
        self.assertIn("--network-check", build_script)
        self.assertIn('collect_all("certifi")', spec)
        self.assertIn("BUNDLE(", spec)

    def test_mac_timer_fallback_is_translated_in_every_language(self):
        expected_keys = {
            "mac_timer_obs_button",
            "mac_timer_obs_title",
            "mac_timer_health_title",
            "mac_timer_health_detail",
            "mac_game_timer_button",
            "mac_level_timer_button",
        }
        for language in ("en", "au", "es", "fr", "de", "pt-BR"):
            with self.subTest(language=language):
                translations = tracker.SETUP_GUIDE_TRANSLATIONS[language]
                self.assertTrue(expected_keys.issubset(translations))
                self.assertTrue(
                    all(translations[key].strip() for key in expected_keys)
                )

    def test_mac_retroarch_switches_safely_then_launches_the_app(self):
        app = tracker.TrackerApp.__new__(tracker.TrackerApp)
        with tempfile.TemporaryDirectory() as temporary_directory:
            root = Path(temporary_directory)
            executable = (
                root
                / "RetroArch.app"
                / "Contents"
                / "MacOS"
                / "RetroArch"
            )
            core = root / "cores" / "snes9x_libretro.dylib"
            rom = root / "roms" / "Test Hack.sfc"
            executable.parent.mkdir(parents=True)
            core.parent.mkdir(parents=True)
            rom.parent.mkdir(parents=True)
            executable.write_bytes(b"APP")
            core.write_bytes(b"CORE")
            rom.write_bytes(b"ROM")

            app.config = {
                "retroarch_executable_path": str(executable),
                "retroarch_core_path": str(core),
            }
            app._resolve_local_rom_path = mock.Mock(
                return_value=(rom, "exact local filename match")
            )
            app._load_retroarch_content_in_place = mock.Mock()
            app._prepare_retroarch_game_switch = mock.Mock(
                return_value="saved previous state and restarted RetroArch"
            )

            with (
                mock.patch.object(tracker, "IS_WINDOWS", False),
                mock.patch.object(tracker, "IS_MACOS", True),
                mock.patch.object(
                    tracker,
                    "launch_local_application",
                ) as launch,
            ):
                expected_config_path = tracker.retroarch_config_path(
                    executable
                )
                result = app._run_local_emulator_launcher(
                    {"title": "Test Hack"},
                    "RetroArch",
                )

            app._load_retroarch_content_in_place.assert_not_called()
            app._prepare_retroarch_game_switch.assert_called_once_with(
                already_saved=False
            )
            launch.assert_called_once_with(
                executable,
                [
                    "--config",
                    str(expected_config_path),
                    "-L",
                    str(core),
                    str(rom),
                ],
            )
            self.assertIn("restarted RetroArch", result["method"])

    def test_mac_retroarch_setup_waits_for_a_game_before_launching(self):
        app = tracker.TrackerApp.__new__(tracker.TrackerApp)
        app.config = {}
        app.root = mock.Mock()
        app.status_var = mock.Mock()
        app._guided_optional_software_completed = mock.Mock()
        dialog = mock.Mock()
        executable = Path(
            "/Users/mario/Applications/RetroArch.app/Contents/MacOS/RetroArch"
        )
        core = Path(
            "/Users/mario/.config/retroarch/cores/"
            "bsnes_mercury_performance_libretro.dylib"
        )

        with (
            mock.patch.object(tracker, "IS_MACOS", True),
            mock.patch.object(tracker, "save_config"),
            mock.patch.object(tracker.messagebox, "showinfo"),
            mock.patch.object(tracker, "launch_local_application") as launch,
        ):
            app._finish_optional_software_install(
                "retroarch",
                executable,
                core,
                dialog,
                "",
            )
        launch.assert_not_called()

    def test_portable_workbook_writer_updates_tracker_sheet(self):
        with tempfile.TemporaryDirectory() as temporary_directory:
            workbook_path = Path(temporary_directory) / "tracker.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = tracker.TRACKER_SHEET
            headers = (
                "Select ROM Hack",
                "Date Started",
                "Completed Exits",
                "Status",
                "Date Completed",
                "Rating (1-5)",
                "Playtime",
                "Minutes",
                "Seconds",
                "Notes",
            )
            sheet.append(headers)
            sheet.append([None] * len(headers))
            sheet.append([None] * len(headers))
            workbook.save(workbook_path)
            workbook.close()

            result = tracker.TrackerWorker._update_tracker_spreadsheet_portable(
                SimpleNamespace(),
                workbook_path,
                "finish",
                "Test Hack",
                5,
                5,
                3723,
                4.5,
                "Mac test",
            )
            self.assertTrue(result["saved"])
            self.assertTrue(result["verified"])
            self.assertEqual(result["percentage"], 100)

            updated = load_workbook(workbook_path, data_only=False)
            updated_sheet = updated[tracker.TRACKER_SHEET]
            self.assertEqual(updated_sheet["A2"].value, "Test Hack")
            self.assertEqual(updated_sheet["C2"].value, "5/5")
            self.assertEqual(updated_sheet["F2"].value, 4.5)
            self.assertEqual(updated_sheet["G2"].value, "1 Hour")
            self.assertEqual(updated_sheet["H2"].value, "2 Minutes")
            self.assertEqual(updated_sheet["I2"].value, "3 Seconds")
            self.assertEqual(updated_sheet["J2"].value, "Mac test")
            updated.close()


if __name__ == "__main__":
    unittest.main()
